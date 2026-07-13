from __future__ import annotations

from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from openpyxl import load_workbook

from agent.graph import _fallback_intent
from agent.controller import AgentSession, InspectionAgent, MAX_MEMORY_MESSAGES
from agent.api_validation import api_key_fingerprint, validate_api_key
from services.errors import explain_exception
from services.files import discover_files, infer_pharmacy_name, unique_output_path
from services.knowledge import KnowledgeBase, load_welcome_message
from services.settings import load_saved_api_key, save_api_key
from pharmacy_analysis import (
    COLUMN_CONFIG_SHEET,
    COLUMN_CONFIG_USER_COLUMN,
    create_column_config_template,
)


class AgentCoreTests(unittest.TestCase):
    def test_api_key_validation_success_and_failure(self) -> None:
        class WorkingModel:
            def invoke(self, _messages):
                return object()

        class InvalidKeyModel:
            def invoke(self, _messages):
                raise RuntimeError("401 invalid api key")

        self.assertTrue(validate_api_key("valid-key", WorkingModel()).valid)
        invalid = validate_api_key("wrong-key", InvalidKeyModel())
        self.assertFalse(invalid.valid)
        self.assertEqual(invalid.code, "AUTH_ERROR")
        self.assertNotEqual(api_key_fingerprint("key-a"), api_key_fingerprint("key-b"))

    def test_api_key_json_round_trip(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "deepseek_config.json"
            save_api_key(path, "secret-key")
            loaded, error = load_saved_api_key(path)
            self.assertIsNone(error)
            self.assertEqual(loaded, "secret-key")
            self.assertNotIn("secret-key", path.name)

    def test_invalid_api_key_json_is_readable_error(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "deepseek_config.json"
            path.write_text("not-json", encoding="utf-8")
            loaded, error = load_saved_api_key(path)
            self.assertIsNone(loaded)
            self.assertIn("无法读取", error) # type: ignore

    def test_fallback_intents(self) -> None:
        self.assertEqual(_fallback_intent("我要做销售医保比对"), "sales_medical")
        self.assertEqual(_fallback_intent("检查处方和医保"), "prescription_medical")
        self.assertEqual(_fallback_intent("确认开始"), "confirm")
        self.assertEqual(_fallback_intent("列名配置里面有什么"), "knowledge")
        self.assertEqual(_fallback_intent("我的列名配置修改好了"), "config_ready")
        self.assertEqual(_fallback_intent("直接使用"), "config_ready")
        self.assertEqual(_fallback_intent("我已经拷贝了相关数据"), "list_files")

    def test_welcome_message_uses_chabao_role(self) -> None:
        root = Path(__file__).resolve().parents[1]
        welcome = load_welcome_message(root / "knowledge")
        self.assertIn("查宝（药店版）", welcome)
        self.assertIn("唐苑清@DriftMousie", welcome)
        self.assertIn("Firefly_bit", welcome)
        self.assertNotIn("唐苑清@漂流小鼠", welcome)

    def test_app_branding_and_avatar_resource(self) -> None:
        root = Path(__file__).resolve().parents[1]
        app_source = (root / "app.py").read_text(encoding="utf-8")
        self.assertIn('APP_TITLE = "查宝（药店版）"', app_source)
        self.assertIn('APP_SUBTITLE = "designed by DriftMousie"', app_source)
        self.assertIn('avatar=avatar', app_source)
        self.assertIn('CHABAO_WORKSPACE', app_source)
        self.assertIn('打开工作目录', app_source)
        self.assertIn('is_processing', app_source)
        self.assertIn('正在执行数据处理，请不要关闭窗口', app_source)
        self.assertIn('disabled=not api_key_is_valid or is_processing', app_source)
        self.assertTrue((root / "resources" / "image.jpg").exists())

    def test_launcher_passes_exe_directory_as_workspace(self) -> None:
        root = Path(__file__).resolve().parents[1]
        launcher_source = (root / "launcher.py").read_text(encoding="utf-8")
        self.assertIn('def _workspace_root()', launcher_source)
        self.assertIn('Path(sys.executable).resolve().parent', launcher_source)
        self.assertIn('CHABAO_WORKSPACE', launcher_source)

    def test_completed_operations_have_knowledge_based_next_steps(self) -> None:
        root = Path(__file__).resolve().parents[1]
        agent = InspectionAgent(root, "test-key", AgentSession())
        config_step = agent._next_step_for("create_config")
        return_step = agent._next_step_for("return_offset")
        sales_step = agent._next_step_for("sales_medical")
        prescription_step = agent._next_step_for("prescription_medical")
        self.assertIn("用户录入列名", config_step)
        self.assertIn("冲销编号", return_step)
        self.assertIn("手工操作", return_step)
        self.assertIn("未匹配", sales_step)
        self.assertIn("先药后方", prescription_step)

    def test_session_remembers_only_latest_ten_rounds(self) -> None:
        session = AgentSession()
        for index in range(1, 12):
            session.remember(f"用户消息{index}", f"查宝回复{index}")
        self.assertEqual(len(session.messages), MAX_MEMORY_MESSAGES)
        history = session.history_text()
        self.assertNotIn("用户：用户消息1\n", history)
        self.assertIn("用户消息2", history)
        self.assertIn("查宝回复11", history)

    def test_knowledge_base_returns_column_config_content(self) -> None:
        root = Path(__file__).resolve().parents[1]
        knowledge = KnowledgeBase(root / "knowledge")
        context = knowledge.context("列名配置里面有什么")
        self.assertIn("字段代码", context)
        self.assertIn("用户录入列名", context)

    def test_agent_question_uses_knowledge_instead_of_static_help(self) -> None:
        root = Path(__file__).resolve().parents[1]
        agent = InspectionAgent(root, "test-key", AgentSession())

        class OfflineModel:
            def invoke(self, _messages):
                raise ConnectionError("offline")

        agent.model = OfflineModel() # type: ignore
        reply, result = agent.handle("列名配置里面有什么")
        self.assertIsNone(result)
        self.assertIn("字段代码", reply)
        self.assertIn("下一步可输入", reply)
        self.assertNotIn("请直接说出需要进行的检查", reply)

    def test_processing_lock_blocks_duplicate_requests(self) -> None:
        root = Path(__file__).resolve().parents[1]
        session = AgentSession(is_processing=True)
        agent = InspectionAgent(root, "test-key", session)
        reply, result = agent.handle("生成列名配置")
        self.assertIsNone(result)
        self.assertIn("当前已有任务正在执行", reply)

    def test_confirm_executes_pending_config_without_model_reclassification(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            agent = InspectionAgent(root, "test-key", AgentSession())

            class FailingGraph:
                def invoke(self, _state):
                    raise AssertionError("明确指令不应交给模型重新判断")

            agent.graph = FailingGraph() # type: ignore
            first_reply, first_result = agent.handle("生成列名配置")
            self.assertIsNone(first_result)
            self.assertIn("回复“确认”", first_reply)
            self.assertIsNotNone(agent.session.pending_action)

            second_reply, second_result = agent.handle("确认")
            self.assertIsNotNone(second_result)
            self.assertTrue(second_result.success) # type: ignore
            self.assertTrue((root / "列名配置.xlsx").exists())
            self.assertIn("列名配置表已生成", second_reply)
            self.assertNotIn("当前目录没有列名配置", second_reply)

    def test_cancel_clears_pending_action_without_model_reclassification(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            session = AgentSession(
                pending_action=None,
            )
            agent = InspectionAgent(root, "test-key", session)
            agent._prepare_config()
            self.assertIsNotNone(agent.session.pending_action)

            class FailingGraph:
                def invoke(self, _state):
                    raise AssertionError("取消不应交给模型重新判断")

            agent.graph = FailingGraph() # type: ignore
            reply, result = agent.handle("取消")
            self.assertIsNone(result)
            self.assertIsNone(agent.session.pending_action)
            self.assertIn("已取消", reply)

    def test_use_existing_config_is_handled_without_model_reclassification(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            create_column_config_template(root / "列名配置.xlsx")
            agent = InspectionAgent(root, "test-key", AgentSession())

            class FailingGraph:
                def invoke(self, _state):
                    raise AssertionError("直接使用配置不应交给模型重新判断")

            agent.graph = FailingGraph() # type: ignore
            reply, result = agent.handle("直接使用")
            self.assertIsNone(result)
            self.assertIn("已经识别到并读取列名配置", reply)

    def test_config_ready_suggests_return_check_not_config_ready_again(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            create_column_config_template(root / "列名配置.xlsx")
            (root / "平民药房销售端.xlsx").touch()
            agent = InspectionAgent(root, "test-key", AgentSession(last_operation="create_config"))

            class FailingGraph:
                def invoke(self, _state):
                    raise AssertionError("配置已完成不应交给模型重新判断")

            agent.graph = FailingGraph() # type: ignore
            reply, result = agent.handle("我的列名配置修改好了")
            self.assertIsNone(result)
            self.assertEqual(agent.session.last_operation, "config_ready")
            self.assertIn("建议下一步执行退货冲销检查", reply)
            self.assertIn("下一步可输入：“执行退货冲销检查”", reply)
            self.assertNotIn("下一步可输入：“我的列名配置修改好了”", reply)

    def test_return_check_without_medical_does_not_suggest_return_check_again(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            create_column_config_template(root / "列名配置.xlsx")
            sales = root / "平民药房销售端.xlsx"
            sales.touch()
            return_output = root / "平民药房_退货冲销检查_2026070912.xlsx"
            return_output.touch()
            session = AgentSession(last_operation="return_offset", return_output=return_output)
            agent = InspectionAgent(root, "test-key", session)
            suggestion = agent._next_input_suggestion()
            guidance = agent._contextual_guidance()
            self.assertIn("查看当前文件", suggestion)
            self.assertIn("医保端文件", guidance)
            self.assertNotIn("执行退货冲销检查", suggestion)

    def test_next_input_suggestion_uses_files_and_last_operation(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "平民药房医保端.csv").touch()
            (root / "平民药房处方明细.xlsx").touch()
            (root / "处方药目录.xlsx").touch()
            create_column_config_template(root / "列名配置.xlsx")
            session = AgentSession(last_operation="sales_medical")
            agent = InspectionAgent(root, "test-key", session)
            self.assertIn("进行处方医保比对", agent._next_input_suggestion())

    def test_discover_files_classifies_known_roles(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            for name in (
                "平民药房销售端.xlsx",
                "平民药房医保端.csv",
                "平民药房处方明细.xlsx",
                "列名配置.xlsx",
                "平民药房_医保销售比对_2026070814.xlsx",
            ):
                (root / name).touch()
            (root / "resources").mkdir()
            (root / "resources" / "处方药目录.xlsx").touch()
            inventory = discover_files(root)
            self.assertEqual([path.name for path in inventory.sales], ["平民药房销售端.xlsx"])
            self.assertEqual([path.name for path in inventory.medical], ["平民药房医保端.csv"])
            self.assertEqual([path.name for path in inventory.prescription], ["平民药房处方明细.xlsx"])
            self.assertEqual([path.name for path in inventory.catalog], ["处方药目录.xlsx"])
            self.assertEqual([path.name for path in inventory.configs], ["列名配置.xlsx"])

    def test_discover_files_tracks_completed_results_and_legacy_excel(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            for name in (
                "常春堂_退货冲销检查_2026071110.xlsx",
                "常春堂_医保销售比对_2026071110.xlsx",
                "常春堂_医保处方比对_2026071111.csv",
                "常春堂销售端.xls",
            ):
                (root / name).touch()
            inventory = discover_files(root)
            self.assertEqual(len(inventory.return_results), 1)
            self.assertEqual(len(inventory.sales_medical_results), 1)
            self.assertEqual(len(inventory.prescription_medical_results), 1)
            self.assertEqual([path.name for path in inventory.unsupported_spreadsheets], ["常春堂销售端.xls"])
            self.assertEqual(inventory.sales, [])

    def test_all_comparisons_completed_suggests_review_not_rerun(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "常春堂_医保销售比对_2026071110.xlsx").touch()
            (root / "常春堂_医保处方比对_2026071111.csv").touch()
            agent = InspectionAgent(root, "test-key", AgentSession(last_operation="prescription_medical"))
            reply = agent._append_next_input_suggestion("处方端医保端综合比对完成。")
            self.assertIn("销售医保比对和处方医保比对均已完成", reply)
            self.assertIn("查看当前文件", reply)
            self.assertNotIn("下一步可输入：“进行销售医保比对”", reply)

    def test_legacy_xls_is_reported_when_expected_file_is_missing(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "常春堂销售端.xls").touch()
            create_column_config_template(root / "列名配置.xlsx")
            agent = InspectionAgent(root, "test-key", AgentSession())
            reply = agent._prepare_return_check()
            self.assertIn("常春堂销售端.xls", reply)
            self.assertIn("另存为 `.xlsx`", reply)

    def test_sales_compare_reminds_manual_offset_cleanup(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            return_output = root / "常春堂_退货冲销检查_2026071110.xlsx"
            return_output.touch()
            (root / "常春堂医保端.csv").touch()
            create_column_config_template(root / "列名配置.xlsx")
            session = AgentSession(return_output=return_output, last_operation="return_offset")
            agent = InspectionAgent(root, "test-key", session)
            reply = agent._prepare_sales_compare()
            self.assertIn("手工删除", reply)
            self.assertIn("冲销编号", reply)

    def test_app_has_stale_processing_lock_recovery(self) -> None:
        root = Path(__file__).resolve().parents[1]
        app_source = (root / "app.py").read_text(encoding="utf-8")
        self.assertIn("def _recover_stale_processing_state()", app_source)
        self.assertIn("session.is_processing = False", app_source)
        self.assertIn('st.session_state.queued_submission = {"user_text": user_text}', app_source)

    def test_output_path_never_overwrites(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            now = datetime(2026, 7, 8, 14, 35)
            first = unique_output_path(root, "平民药房", "医保销售比对", ".xlsx", now)
            first.touch()
            second = unique_output_path(root, "平民药房", "医保销售比对", ".xlsx", now)
            self.assertEqual(first.name, "平民药房_医保销售比对_2026070814.xlsx")
            self.assertEqual(second.name, "平民药房_医保销售比对_2026070814_2.xlsx")

    def test_pharmacy_name_fallback(self) -> None:
        self.assertEqual(infer_pharmacy_name(Path("平民健康药房销售端.xlsx")), "平民健康药房")
        self.assertEqual(infer_pharmacy_name(Path("销售端.xlsx")), "XX药店")

    def test_config_template_highlights_user_column_header(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "列名配置.xlsx"
            create_column_config_template(path)
            worksheet = load_workbook(path)[COLUMN_CONFIG_SHEET]
            headers = [cell.value for cell in worksheet[1]]
            cell = worksheet.cell(1, headers.index(COLUMN_CONFIG_USER_COLUMN) + 1)
            self.assertTrue(cell.font.bold)
            self.assertEqual(cell.font.color.rgb, "FFFF0000")

    def test_sales_compare_can_warn_and_continue_without_return_check(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "平民药房销售端.xlsx").touch()
            (root / "平民药房医保端.csv").touch()
            create_column_config_template(root / "列名配置.xlsx")
            agent = InspectionAgent(root, "test-key", AgentSession())
            reply = agent._prepare_sales_compare()
            self.assertIn("还没有在本次会话中完成退货冲销", reply)
            self.assertIsNotNone(agent.session.pending_action)
            self.assertEqual(agent.session.pending_action.name, "sales_medical") # type: ignore

    def test_missing_column_error_is_readable(self) -> None:
        error = explain_exception(ValueError("输入表缺少必要列：流水号"), "销售医保比对")
        self.assertEqual(error.error_code, "MISSING_REQUIRED_COLUMN")
        self.assertIn("流水号", error.user_message)


if __name__ == "__main__":
    unittest.main()
